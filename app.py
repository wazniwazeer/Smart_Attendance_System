import json, sqlite3, csv, time
from pathlib import Path
from datetime import datetime

import numpy as np
import cv2
import torch
import torch.nn as nn
from flask import Flask, request, jsonify, Response

# Optional YOLO (face detector)
try:
    from ultralytics import YOLO
except Exception:
    YOLO = None

# ================== Config & Artifacts ==================
ARTIFACTS = Path("artifacts")
ARTIFACTS.mkdir(parents=True, exist_ok=True)

CFG = {
    "yolo_face_weights": "",     # e.g., "artifacts/yolov8n-face.pt" (optional)
    "yolo_conf": 0.35,
    "yolo_iou": 0.50,
    "yolo_imgsz": 640,
    "max_det": 3,
    "margin": 0.30,
    "img_size_for_emb": 160,
    "unknown_conf_thres": 0.55,  # stricter default so unknowns → Face Not Found
    "embed_dim": 512,
    "dwell_seconds": 2.0,        # must be seen this long before logging
    "absent_reset_sec": 5.0      # session resets if not seen for this long
}
cfg_path = ARTIFACTS / "config.json"
if cfg_path.exists():
    try:
        CFG.update(json.loads(cfg_path.read_text()))
    except Exception as e:
        print("[WARN] Failed to parse config.json:", e)

DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
print(f"[INFO] Device: {DEVICE}")

# ================== Preprocess utils ==================
def fixed_image_standardization(t: torch.Tensor) -> torch.Tensor:
    # expects float tensor in [0,1] → returns [-1,1]
    return (t - 0.5) / 0.5

def to_face_tensor(rgb_np: np.ndarray, size: int) -> torch.Tensor:
    img = cv2.resize(rgb_np, (size, size)).astype(np.float32) / 255.0  # [0,1]
    t = torch.from_numpy(img).permute(2, 0, 1)                          # [3,H,W]
    t = fixed_image_standardization(t)                                   # [-1,1]
    return t.float()

def expand_box(box, img_shape, margin=0.30):
    h, w = img_shape[:2]
    x1, y1, x2, y2 = map(int, box)
    bw, bh = x2 - x1, y2 - y1
    x1 = max(0, x1 - int(bw * margin))
    y1 = max(0, y1 - int(bh * margin))
    x2 = min(w, x2 + int(bw * margin))
    y2 = min(h, y2 + int(bh * margin))
    return [x1, y1, x2, y2]

def crop_face_rgb(img_bgr, box, size):
    x1, y1, x2, y2 = map(int, box)
    face = img_bgr[y1:y2, x1:x2]
    if face.size == 0:
        return None
    face = cv2.cvtColor(face, cv2.COLOR_BGR2RGB)
    face = cv2.resize(face, (size, size))
    return face

def too_small(box, min_side=80):
    x1, y1, x2, y2 = map(int, box)
    return (x2 - x1) < min_side or (y2 - y1) < min_side

# ================== Model ==================
from facenet_pytorch import InceptionResnetV1

def build_head(embed_dim: int, num_classes: int) -> nn.Sequential:
    # Plain Sequential so saved state_dict keys ("0.*","2.*") match
    return nn.Sequential(
        nn.LayerNorm(embed_dim),
        nn.Dropout(0.1),
        nn.Linear(embed_dim, num_classes)
    )

class FaceRecognizer:
    def __init__(self):
        # Labels
        labels_path = ARTIFACTS / "label_classes.npy"
        assert labels_path.exists(), f"Missing {labels_path}"
        self.classes = np.load(labels_path, allow_pickle=True)
        self.classes = np.array([str(c).strip() for c in self.classes], dtype=object)
        self.num_classes = len(self.classes)

        # Embedder
        self.embedder = InceptionResnetV1(pretrained='vggface2').eval().to(DEVICE)

        # Head
        embed_dim = int(CFG.get("embed_dim", 512))
        self.head = build_head(embed_dim, self.num_classes).to(DEVICE)

        head_path = ARTIFACTS / "head_best.pth"
        if not head_path.exists():
            alt = ARTIFACTS / "head_best.h5"
            if alt.exists():
                head_path = alt
        assert head_path.exists(), "head_best.pth/.h5 not found in artifacts/"

        sd = torch.load(head_path, map_location=DEVICE)
        try:
            self.head.load_state_dict(sd, strict=True)
        except RuntimeError as e:
            print("[WARN] strict=True failed, loading with strict=False:", e)
            self.head.load_state_dict(sd, strict=False)
        self.head.eval()

        # Optional index remap if your head's class order differs from label_classes
        remap_path = ARTIFACTS / "index_remap.npy"
        if remap_path.exists():
            self.index_remap = np.load(remap_path).astype(int)
            if self.index_remap.shape[0] != self.num_classes:
                raise RuntimeError("index_remap.npy length != num_classes.")
            print("[INFO] Using index remap from artifacts/index_remap.npy")
        else:
            self.index_remap = np.arange(self.num_classes, dtype=int)

        # Detector (optional but recommended)
        self.detector = None
        if YOLO is not None and CFG.get("yolo_face_weights"):
            w = CFG["yolo_face_weights"]
            if Path(w).exists():
                self.detector = YOLO(str(w))
                print("[INFO] YOLO face model loaded:", w)
            else:
                print("[WARN] YOLO weights not found; detector disabled.")
        else:
            print("[INFO] YOLO not used; detector disabled (full frame fallback).")

        self.unknown_th = float(CFG.get("unknown_conf_thres", 0.55))
        self.size = int(CFG.get("img_size_for_emb", 160))

    @torch.inference_mode()
    def _embed_tensor(self, face_tensor: torch.Tensor) -> np.ndarray:
        emb = self.embedder(face_tensor.unsqueeze(0).to(DEVICE))
        return emb.cpu().numpy().reshape(-1)

    def _detect_best_box(self, img_bgr):
        if self.detector is None:
            h, w = img_bgr.shape[:2]
            return [0, 0, w, h], 1.0
        res = self.detector.predict(
            source=img_bgr, verbose=False,
            conf=CFG["yolo_conf"], iou=CFG["yolo_iou"],
            imgsz=CFG["yolo_imgsz"], max_det=CFG["max_det"]
        )
        best = None
        for r in res:
            if r.boxes is None:
                continue
            b_xyxy = r.boxes.xyxy.cpu().numpy()
            b_conf = r.boxes.conf.cpu().numpy()
            if len(b_conf) == 0:
                continue
            j = int(np.argmax(b_conf))
            best = (b_xyxy[j].tolist(), float(b_conf[j]))
        return (best[0], best[1]) if best else (None, 0.0)

    @torch.inference_mode()
    def predict_from_bgr(self, img_bgr):
        # 1) detect
        box, _det_conf = self._detect_best_box(img_bgr)
        if box is None or too_small(box):
            return {"recognized": False, "message": "Face Not Found"}

        # 2) crop + embed
        box = expand_box(box, img_bgr.shape, margin=CFG["margin"])
        face_rgb = crop_face_rgb(img_bgr, box, size=self.size)
        if face_rgb is None:
            return {"recognized": False, "message": "Face Not Found"}

        t = to_face_tensor(face_rgb, self.size)
        emb = self._embed_tensor(t)
        x = torch.from_numpy(emb).float().unsqueeze(0).to(DEVICE)

        # 3) classify
        logits = self.head(x)                                 # [1,C]
        prob = torch.softmax(logits, dim=1).cpu().numpy()[0]  # [C]
        idx_raw = int(np.argmax(prob))
        idx = int(self.index_remap[idx_raw])                  # map logits -> label index
        conf = float(prob[idx_raw])

        # 4) threshold → single clean result
        if conf < self.unknown_th:
            return {"recognized": False, "message": "Face Not Found"}

        name = str(self.classes[idx])
        return {"recognized": True, "name": name, "confidence": conf}

# ================== Attendance storage (DB/CSV/Excel) ==================
DB_PATH = ARTIFACTS / "attendance.db"
CSV_PATH = ARTIFACTS / "attendance.csv"
XLSX_PATH = ARTIFACTS / "attendance.xlsx"
XLSX_SHEET = "Attendance"

def _db_connect():
    return sqlite3.connect(DB_PATH)

def init_db():
    with _db_connect() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            name TEXT NOT NULL,
            date TEXT NOT NULL,
            time TEXT NOT NULL,
            confidence REAL NOT NULL,
            timestamp_iso TEXT NOT NULL,
            PRIMARY KEY (name, date)    -- one entry per person per day
        )
        """)
        con.commit()

def insert_attendance(name: str, confidence: float) -> bool:
    """Insert into SQLite; returns True if a new row was created (not duplicate)."""
    now = datetime.now()
    row = (name, now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"),
           float(confidence), now.isoformat())
    try:
        with _db_connect() as con:
            con.execute("INSERT INTO attendance (name,date,time,confidence,timestamp_iso) VALUES (?,?,?,?,?)", row)
            con.commit()
        # mirrored write to CSV & Excel only on actual insert
        _append_csv(row)
        _append_excel(row)
        return True
    except sqlite3.IntegrityError:
        # already logged today for this name
        return False

def _append_csv(row_tuple):
    is_new = not CSV_PATH.exists()
    with open(CSV_PATH, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if is_new:
            w.writerow(["name", "date", "time", "confidence", "timestamp_iso"])
        w.writerow(row_tuple)

def _append_excel(row_tuple):
    from openpyxl import Workbook, load_workbook
    headers = ["name", "date", "time", "confidence", "timestamp_iso"]
    if not XLSX_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = XLSX_SHEET
        ws.append(headers)
        wb.save(XLSX_PATH)
    wb = load_workbook(XLSX_PATH)
    ws = wb[XLSX_SHEET] if XLSX_SHEET in wb.sheetnames else wb.active
    ws.append(list(row_tuple))
    wb.save(XLSX_PATH)

# ================== Session gate (lock + dwell) ==================
DWELL_SECONDS = float(CFG["dwell_seconds"])
ABSENT_RESET = float(CFG["absent_reset_sec"])
SESSIONS = {}  # key: (name, date) -> {first_seen, last_seen, logged}

def update_session_and_maybe_log(name: str, conf: float, now_ts: float):
    day = datetime.now().strftime("%Y-%m-%d")
    key = (name, day)
    s = SESSIONS.get(key)
    if s is None or (now_ts - s["last_seen"] > ABSENT_RESET):
        SESSIONS[key] = {"first_seen": now_ts, "last_seen": now_ts, "logged": False}
        return False
    s["last_seen"] = now_ts
    if not s["logged"] and (now_ts - s["first_seen"] >= DWELL_SECONDS):
        if insert_attendance(name, conf):
            s["logged"] = True
            return True
    return False

def cleanup_sessions(now_ts: float):
    remove = [k for k, s in SESSIONS.items() if now_ts - s["last_seen"] > ABSENT_RESET]
    for k in remove:
        del SESSIONS[k]

# ================== Flask + Realtime UI ==================
app = Flask(__name__)
init_db()
rec = FaceRecognizer()

@app.get("/")
def index():
    # Simple UI: webcam -> JPEG -> /api/predict
    html = """
<!doctype html>
<html>
<head>
  <meta charset="utf-8" />
  <title>Smart AI Attendance — Realtime</title>
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <style>
    body { font-family: system-ui, -apple-system, Segoe UI, Roboto, Arial; background:#0b0d10; color:#eaecef; }
    .wrap { max-width: 900px; margin: 24px auto; padding: 0 16px; }
    .card { background:#111418; border:1px solid #202530; border-radius:12px; padding:16px; box-shadow:0 2px 20px rgba(0,0,0,.25); }
    video, canvas { width:100%; max-width:640px; border-radius:12px; border:1px solid #202530; }
    .badge { display:inline-block; padding:6px 10px; border-radius:999px; background:#1f6feb; color:#fff; font-size:14px; }
    .muted { color:#93a1ad; font-size:14px; }
    #status { margin-top:12px; }
  </style>
</head>
<body>
<div class="wrap">
  <h2>👤 Smart AI Face-Recognition — Realtime</h2>
  <div class="card">
    <video id="video" autoplay playsinline muted></video>
    <canvas id="canvas" style="display:none;"></canvas>
    <div id="status" class="badge">Initializing…</div>
    <div style="margin-top:10px;">
      <label>Interval (ms):
        <input type="number" id="ival" value="300" min="100" step="50" style="width:100px;">
      </label>
      <button id="btnStart">Start</button>
      <button id="btnStop">Stop</button>
    </div>
    <div id="result" class="muted" style="margin-top:10px; font-size:18px;">—</div>
  </div>
</div>

<script>
const video = document.getElementById('video');
const canvas = document.getElementById('canvas');
const statusEl = document.getElementById('status');
const resultEl = document.getElementById('result');
const ivalEl = document.getElementById('ival');
const btnStart = document.getElementById('btnStart');
const btnStop  = document.getElementById('btnStop');

let timer = null;

async function initCam() {
  try {
    const stream = await navigator.mediaDevices.getUserMedia({ video: true, audio: false });
    video.srcObject = stream;
    statusEl.textContent = "Camera ready";
  } catch (e) {
    statusEl.textContent = "Camera error: " + e.message;
  }
}

function snapAndSend() {
  if (!video.videoWidth) return;
  canvas.width = video.videoWidth; canvas.height = video.videoHeight;
  const ctx = canvas.getContext('2d');
  ctx.drawImage(video, 0, 0, canvas.width, canvas.height);
  canvas.toBlob(async (blob) => {
    if (!blob) return;
    const fd = new FormData();
    fd.append('file', blob, 'frame.jpg');
    try {
      const res = await fetch('/api/predict', { method: 'POST', body: fd });
      const json = await res.json();
      if (!res.ok) throw new Error(JSON.stringify(json));
      if (json.recognized) {
        const conf = (json.confidence ?? 0).toFixed(3);
        // If recorded==true, show a checkmark once
        resultEl.textContent = `Name: ${json.name}  •  Confidence: ${conf}` + (json.recorded ? "  ✓ recorded" : "");
      } else {
        resultEl.textContent = "Face Not Found";
      }
      statusEl.textContent = "Streaming…";
    } catch (e) {
      statusEl.textContent = "Error";
      resultEl.textContent = "Error";
      console.error(e);
    }
  }, 'image/jpeg', 0.85);
}

btnStart.onclick = () => {
  if (timer) return;
  const ms = Math.max(100, parseInt(ivalEl.value || "300", 10));
  timer = setInterval(snapAndSend, ms);
  statusEl.textContent = "Streaming…";
};
btnStop.onclick = () => {
  if (timer) { clearInterval(timer); timer = null; }
  statusEl.textContent = "Stopped";
};

initCam();
</script>
</body>
</html>
"""
    return Response(html, mimetype="text/html")

@app.post("/api/predict")
def api_predict():
    if "file" not in request.files:
        return jsonify({"error": "no file"}), 400
    data = request.files["file"].read()
    arr = np.frombuffer(data, np.uint8)
    bgr = cv2.imdecode(arr, cv2.IMREAD_COLOR)
    if bgr is None:
        return jsonify({"error": "decode_failed"}), 400

    out = rec.predict_from_bgr(bgr)
    now_ts = time.time()

    recorded = False
    if out.get("recognized"):
        # Lock + dwell + dedup (one row per name/day)
        recorded = update_session_and_maybe_log(out["name"], out["confidence"], now_ts)
    cleanup_sessions(now_ts)

    # Return single clean result with "recorded" flag
    if out.get("recognized"):
        return jsonify({"recognized": True, "name": out["name"], "confidence": out["confidence"], "recorded": bool(recorded)})
    else:
        return jsonify({"recognized": False, "message": "Face Not Found", "recorded": False})

@app.get("/health")
def health():
    return jsonify({"status": "ok", "device": "cuda" if torch.cuda.is_available() else "cpu",
                    "time": datetime.now().isoformat()})

if __name__ == "__main__":
    # Local dev server
    app.run(host="0.0.0.0", port=5000, debug=True)
